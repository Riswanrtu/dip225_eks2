from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    priority = row[7]       # Column H (index 7)
    delivery_date = row[9]  # Column J (index 9)

    if priority == 'High' and hasattr(delivery_date, 'year') and delivery_date.year == 2015:
        count += 1

print(count)
