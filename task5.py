from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx', data_only=True)
ws = wb['Lapa_0']

total_sum = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    client = row[5]    # Column F
    quantity = row[11] # Column L
    total = row[13]    # Column N

    if client != 'Korporatīvais' or quantity is None or total is None:
        continue

    try:
        quantity = float(quantity)
        if 40 <= quantity <= 50:
            total_sum += float(total)
    except ValueError:
        continue

print(int(total_sum))
