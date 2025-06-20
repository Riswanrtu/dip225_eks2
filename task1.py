from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    address, number = row[3], row[11]  # Column D is index 3, L is index 11

    if isinstance(address, str) and address.startswith('Ain') and isinstance(number, (int, float)) and number < 40:
        count += 1

print(count)
