from openpyxl import load_workbook

wb = load_workbook('sagatave_eksamenam.xlsx')
ws = wb['Lapa_0']

count = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    address = row[3]  # Column D
    city = row[4]     # Column E

    if isinstance(address, str) and 'Adulienas iela' in address and city in ['Valmiera', 'Saulkrasti']:
        count += 1

print(count)
